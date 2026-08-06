import openpyxl
import pandas as pd
import streamlit as st

from database import get_expenses

from database import get_expenses

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference
from datetime import datetime
from openpyxl.styles import Font, Alignment
from openpyxl.chart import PieChart, Reference

def get_expense_dataframe(database_name):

    expenses = get_expenses(
        database_name
    )

    df = pd.DataFrame(
        expenses,
        columns=[
            "ID",
            "Date",
            "Merchant",
            "Category",
            "Amount",
            "Description",
            "Payment Method",
            "Receipt Image",
            "Created At"
        ]
    )

    return df   # <-- IMPORTANT

def create_summary_sheet(wb, df, selected_month):

    summary = wb.create_sheet("Summary")

    # Title
    summary["A1"] = "Expense Tracker Pro"
    summary["A1"].font = Font(
        bold=True,
        size=18
)

    summary["A2"] = f"{selected_month} Expense Report"
    summary["A2"].font = Font(
        bold=True,
        size=14
)

    # KPI calculations
    total_spending = df["Amount"].sum()
    number_expenses = len(df)
    average_expense = df["Amount"].mean()
    highest_expense = df["Amount"].max()
    lowest_expense = df["Amount"].min()


    # KPI labels
    summary["A4"] = "Metric"
    summary["B4"] = "Value"

    summary["A5"] = "Total Spending"
    summary["B5"] = total_spending

    summary["A6"] = "Number of Expenses"
    summary["B6"] = number_expenses

    summary["A7"] = "Average Expense"
    summary["B7"] = average_expense

    summary["A8"] = "Highest Expense"
    summary["B8"] = highest_expense

    summary["A9"] = "Lowest Expense"
    summary["B9"] = lowest_expense


    # Format headers
    summary["A4"].font = Font(bold=True)
    summary["B4"].font = Font(bold=True)


    # Currency format
    for cell in [
        "B5",
        "B7",
        "B8",
        "B9"
    ]:
        summary[cell].number_format = '€#,##0.00'


    # Column width
    summary.column_dimensions["A"].width = 25
    summary.column_dimensions["B"].width = 18


    # Alignment
    for row in summary.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                horizontal="left"
            )


def create_charts_sheet(wb, df):

    charts = wb.create_sheet("Charts")

    # Prepare category data
    category_data = (
        df.groupby("Category")["Amount"]
        .sum()
    )

    charts["A1"] = "Category"
    charts["B1"] = "Amount"


    row = 2

    for category, amount in category_data.items():

        charts.cell(row=row, column=1).value = category
        charts.cell(row=row, column=2).value = amount

        row += 1


    # Create pie chart
    chart = PieChart()

    labels = Reference(
        charts,
        min_col=1,
        min_row=2,
        max_row=row-1
    )

    values = Reference(
        charts,
        min_col=2,
        min_row=1,
        max_row=row-1
    )


    chart.add_data(
        values,
        titles_from_data=True
    )

    chart.set_categories(labels)

    chart.title = "Expenses by Category"


    charts.add_chart(
        chart,
        "D2"
    )

def export_excel(df, selected_month):

    filename = "Expense_Tracker_Pro_Report.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expenses"


    # Title
    ws["A1"] = "Expense Tracker Pro"
    ws["A1"].font = Font(
        bold=True,
        size=18
    )

    ws["A2"] = "Expense Report"
    ws["A3"] = f"Generated: {datetime.today().strftime('%d/%m/%Y')}"


    start_row = 5


    # Headers
    for col, name in enumerate(df.columns, 1):
        cell = ws.cell(
            row=start_row,
            column=col,
            value=name
        )

        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            "solid",
            fgColor="4F81BD"
        )
        cell.alignment = Alignment(
            horizontal="center"
        )


    # Data
    for row in df.itertuples(index=False):

        ws.append(row)


    # Currency format
    amount_column = None

    for cell in ws[start_row]:

        if cell.value == "Amount":
            amount_column = cell.column


    if amount_column:

        for row in range(start_row + 1, ws.max_row + 1):

            ws.cell(
                row,
                amount_column
            ).number_format = '€#,##0.00'


    # Date format
    for row in range(start_row + 1, ws.max_row + 1):

        ws.cell(row,2).number_format = "DD/MM/YYYY"



    # Total row
    total_row = ws.max_row + 2

    ws.cell(
        total_row,
        4,
        "TOTAL SPENDING"
    ).font = Font(bold=True)

    ws.cell(
        total_row,
        5,
        f"=SUM(E{start_row+1}:E{total_row-2})"
    ).font = Font(bold=True)

    ws.cell(
        total_row,
        5
    ).number_format = '€#,##0.00'


    # Filter
    ws.auto_filter.ref = (
        f"A{start_row}:"
        f"{get_column_letter(ws.max_column)}{ws.max_row}"
    )


    # Freeze header
    ws.freeze_panes = "A6"


    # Auto width
    for column in ws.columns:

        max_length = 0
        letter = get_column_letter(column[0].column)

        for cell in column:

            if cell.value:
                max_length = max(
                    max_length,
                    len(str(cell.value))
                )

        ws.column_dimensions[letter].width = max_length + 3

#  Add Summary and Charts
    create_summary_sheet(wb, df, selected_month)

    create_charts_sheet(wb, df)

    wb.save(filename)

    return filename



def export_pdf(df, selected_month):

    from reportlab.platypus import SimpleDocTemplate, Table, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet
    import pandas as pd

    filename = "Expense_Tracker_Pro_Report.pdf"

    # Remove receipt image from PDF
    if "Receipt Image" in df.columns:
        df = df.drop(columns=["Receipt Image"])

    # Remove time from date
    if "Date" in df.columns:
        df["Date"] = pd.to_datetime(
            df["Date"]
        ).dt.strftime("%d/%m/%Y")


    doc = SimpleDocTemplate(filename)

    styles = getSampleStyleSheet()

    elements = []

    elements.append(
        Paragraph(
            "Expense Tracker Pro Report",
            styles["Title"]
        )
    )


    data = [
        list(df.columns)
    ]

    for row in df.values:
        data.append(list(row))


    table = Table(data)

    elements.append(table)

    doc.build(elements)

    return filename